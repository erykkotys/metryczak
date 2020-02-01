import sqlite3
import re
import os
import logging
import openpyxl
from openpyxl.styles import NamedStyle, Font

print('Metryczak-Beta v 1.0')

# TODO: Funkcja importowania do bazy z xlsx ew. z skanowania folderow (porownywanie z baza www?)
# TODO: MENU do obslugi
# TODO: Obsluga wielu plikow na raz
# TODO: Plik ustawien

# Basic logging settings
logging.basicConfig(level=logging.WARNING, format='%(asctime)s %(levelname)s %(message)s')

# Opening a connection to song database and creating a 'cursor' to execute SQL commands
sql_connection = sqlite3.connect('songs.db')
sql = sql_connection.cursor()

# Input path is received form command line as argument
while True:
    print('Podaj nazwe pliku csv:')
    input_file = input()
    if input_file == '':
        input_file = 'wk 314.csv'
    try:
        with open(input_file, 'r') as file:
            input_content = file.readlines()
            break
    except FileNotFoundError:
        logging.error('Podany plik nie istnieje')
        continue

# Output path is same
output_file = os.path.splitext(input_file)[0] + '.xlsx'
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except PermissionError:
        logging.error(f'Nie moge utworzyc pliku {output_file}, plik istnieje i jest uzywany przez inna aplikacje.\n'
              f'Zamknij plik i sprobuj ponownie')
        exit()

current_row = 1


#  Adding new entries to database
def db_add():
    print(f'Dodaje nowy wpis do bazy dla {line}')
    print('Podaj tytuł utworu:')
    title = input()
    print('Podaj kompozytorów utworu:')
    composers = input()
    with sql_connection:
        sql.execute("INSERT INTO songs VALUES (:title, :composers)", {'title': title, 'composers': composers})
    sheet_append(title, composers, time)


# Deleting existing entries in database
def db_del(title, composers):
    with sql_connection:
        sql.execute("DELETE FROM songs WHERE title = :title AND composers = :composers",
                    {'title': title, 'composers': composers})
    logging.debug('Rejestr usuniety poprawnie')


def sheet_append(title, composers, time):
    global current_row
    sheet.cell(row=current_row, column=1).value = title
    sheet.cell(row=current_row, column=2).value = composers
    sheet.cell(row=current_row, column=3).value = time
    logging.info(f'Zapisuje do pliku: {title}, {composers}, {time}\n')
    current_row += 1


def intro_check():
    if line[:2] == 'WK' or line[:2] == 'SW':
        logging.info(f'Zidentyfikowalem wiersz jako element oprawy WK')
        sheet_append('Time Waits', 'Naylor Matthew Todd, Spencer Oliver', time)
    if line.startswith('Lombard') or line.startswith('JING'):
        logging.info(f'Zidentyfikowalem wiersz jako element oprawy Lombardu')
        sheet_append('Wrap\'Em Up', 'Marlow Mitchell', time)
    if line.startswith('REP'):
        logging.info(f'Zidentyfikowalem wiersz jako element oprawy Reporterow')
        sheet_append('Czolowka', '', time)


def db_match(title_search, char_no):
    if char_no < 2:
        logging.debug('Nie znalazlem utworu w bazie')
        print('Nie znalazlem utworu w bazie, dodaj nowy wpis')
        db_add()
        return -1
    while True:
        with sql_connection:
            matches = sql.execute("SELECT * FROM songs WHERE LOWER(title) LIKE :title",
                                  {'title': title_search[0:char_no].lower()+'%'})
            results = matches.fetchall()
            if len(results) > 1:
                logging.warning(f'Znalazlem {len(results)} rejestrow pasujacych do wpisu {title_search}:')
                for index, result in enumerate(results):
                    print(f'{index}. {result}')
                print(f'{len(results)}. Usun jeden z powyzszych rejestrow z bazy')
                print(f'{len(results)+1}. Dodaj nowy rejestr do bazy i pliku')
                print('Wybierz resjestr do zapisania w pliku lub dodaj nowy (default=0):')
                try:
                    choice = int(input())
                except ValueError:
                    choice = 0
                if choice == len(results):
                    print('Podaj numer rejestru do usuniecia (lub wcisnij Enter by wrocic):')
                    try:
                        delete_choice = int(input())
                    except ValueError:
                        break
                    logging.warning(f'Usuwam z bazy rejestr {results[delete_choice]}')
                    db_del(results[delete_choice][0], results[delete_choice][1])
                    continue
                if choice == len(results)+1:
                    db_add()
                    break
                else:
                    sheet_append(results[choice][0], results[choice][1], time)
                    break
                    # return 0
            if not results:
                db_match(title_search, char_no-1)
                break
            else:
                logging.debug(f'Dopasowalem rejestr z bazy: {results}')
                sheet_append(results[0][0], results[0][1], time)
                break
                return 0


# Start a new Workbook
new_workbook = openpyxl.Workbook()
sheet = new_workbook['Sheet']
sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 50
sheet.column_dimensions['C'].width = 9
header_style = NamedStyle(name='header_style')
header_style.font = Font(bold=True)
new_workbook.add_named_style(header_style)
logging.info('Tworze wiersz naglowka:')
sheet_append('Tytuł', 'Autorzy (Nazwisko, Imię)', 'Czas')
for cell in sheet[1]:
    cell.style = header_style

# Analyse input content
for line in input_content:
    time = None
    logging.debug(f'Analizuje nowy wiersz: {line.rstrip()}')
    time_pattern = re.compile(r'\d\d:\d\d:\d\d')
    try:
        time = time_pattern.search(line).group()
    except AttributeError:
        logging.debug(f'Nie znalazlem czasu utworu w wierszu {line}')
    if time is not None:
        logging.debug(f'Czas utworu {time}')
    intro_check()
    title_pattern = re.compile(r'(\d_[A-Z]\w+_(Main|Instrumental)?)|(---\w+--)')
    title = title_pattern.search(line)
    if title is not None:
        title = title.group(0).strip('1234567890').strip('_-').replace('_', ' ')
        logging.debug(f'Dopasowalem do regexa fragment wiersza: {title}')
        db_match(title, 20)
    else:
        continue


new_workbook.save(output_file)
logging.info(f'Plik {output_file} zapisany poprawnie')
